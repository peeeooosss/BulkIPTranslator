import os
import tkinter as tk
from tkinter import filedialog, messagebox
import requests
from openpyxl import load_workbook, Workbook
import csv
import concurrent.futures
import time

def load_ips(file_path, column, start_row):
    ips = []
    if file_path.endswith('.xlsx'):
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        col_idx = ord(column.upper()) - ord('A')
        for row in sheet.iter_rows(min_row=start_row, min_col=col_idx+1, max_col=col_idx+1, values_only=True):
            if row[0]:
                ips.append(row[0])
    elif file_path.endswith('.csv'):
        with open(file_path, mode='r') as file:
            reader = csv.reader(file)
            for _ in range(start_row - 1):
                next(reader)
            col_idx = ord(column.upper()) - ord('A')
            for row in reader:
                if row[col_idx]:
                    ips.append(row[col_idx])
    return ips

def get_geolocation(ip, api_key):
    url = f"https://ipinfo.io/{ip}/json?token={api_key}"
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        return {
            'IP Address': ip,
            'ISP': data.get('org', 'N/A'),
            'City': data.get('city', 'N/A'),
            'Region': data.get('region', 'N/A'),
            'Postal Code': data.get('postal', 'N/A')
        }
    except requests.exceptions.RequestException:
        return {'IP Address': ip, 'ISP': 'Error', 'City': 'Error', 'Region': 'Error', 'Postal Code': 'Error'}

def write_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "IP Geolocation"

    # Write headers
    headers = ['IP Address', 'ISP', 'City', 'Region', 'Postal Code']
    ws.append(headers)

    # Write data
    for row in data:
        ws.append([row.get(header, '') for header in headers])

    # Adjust the column width for ISP
    ws.column_dimensions['B'].width = 50

    # Ensure the Output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # Save to file
    wb.save(output_file)

def process_file(file_path, column, start_row, api_key):
    try:
        ips = load_ips(file_path, column, start_row)

        with concurrent.futures.ThreadPoolExecutor(max_workers=100) as executor:
            futures = [executor.submit(get_geolocation, ip, api_key) for ip in ips]
            geolocation_data = [future.result() for future in concurrent.futures.as_completed(futures)]

        output_dir = "Output"
        output_file = os.path.join(output_dir, 'ip_geolocation.xlsx')
        write_to_excel(geolocation_data, output_file)
        messagebox.showinfo("Success", f"Geolocation data has been written to {output_file}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")])
    if file_path:
        file_path_entry.delete(0, tk.END)
        file_path_entry.insert(0, file_path)

def on_process():
    file_path = file_path_entry.get().strip()
    column = column_entry.get().strip()
    start_row = int(start_row_entry.get().strip())
    api_key = api_key_entry.get().strip()
    process_file(file_path, column, start_row, api_key)

# Create the GUI application
app = tk.Tk()
app.title("IP Geolocation Processor")

tk.Label(app, text="Select Input File:").grid(row=0, column=0, padx=10, pady=5)
file_path_entry = tk.Entry(app, width=50)
file_path_entry.grid(row=0, column=1, padx=10, pady=5)
tk.Button(app, text="Browse", command=browse_file).grid(row=0, column=2, padx=10, pady=5)

tk.Label(app, text="IP Column (e.g., B):").grid(row=1, column=0, padx=10, pady=5)
column_entry = tk.Entry(app)
column_entry.grid(row=1, column=1, padx=10, pady=5)

tk.Label(app, text="Start Row (e.g., 2):").grid(row=2, column=0, padx=10, pady=5)
start_row_entry = tk.Entry(app)
start_row_entry.grid(row=2, column=1, padx=10, pady=5)

tk.Label(app, text="API Key:").grid(row=3, column=0, padx=10, pady=5)
api_key_entry = tk.Entry(app, width=50)
api_key_entry.grid(row=3, column=1, padx=10, pady=5)

tk.Button(app, text="Process", command=on_process).grid(row=4, column=1, padx=10, pady=20)

app.mainloop()
